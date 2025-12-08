# utils/helpers.py

from .config import SONG_DATABASE_PATH, SORTED_SONG_DATABASE_PATH
import pandas as pd
from pypinyin import pinyin, Style
import math
from openpyxl import load_workbook


# read database.xlsx, concatenate all sheets into a list of structures
def load_song_data():
    # Read all sheets into a dictionary of DataFrames
    all_sheets_dict = pd.read_excel(SONG_DATABASE_PATH, sheet_name=None)

    # # Combine all DataFrames
    # all_songs_df = pd.concat(all_sheets_dict.values(), ignore_index=True)
    #
    # # Remove completely empty rows
    # all_songs_df = all_songs_df.dropna(how="all")
    #
    # # save it
    # all_songs_df.to_excel("output.xlsx", sheet_name='Sheet1', index=False)

    # Loop through the dictionary to access all DataFrames
    songs = []
    for artist, df in all_sheets_dict.items():
        # Remove completely empty rows
        df = df.dropna(how="all")
        for index, row in df.iterrows():
            audio_file = row['File'] + ".mp3" if type(row['File']) == str else ""
            lyrics = row['Lyrics'] if type(row['Lyrics']) == str else ""
            speed = 1
            if 'Speed' in df.columns:
                if type(row['Speed']) == float and not math.isnan(row['Speed']):
                    speed = row['Speed']
            song = {
                'id': index,
                'title': row['Song'],
                'artist': artist,
                'audio_file': audio_file,
                'lyrics': lyrics,
                'type': row['Type'],
                'valid': row['Valid'],
                'speed': speed,
            }
            songs.append(song)
    return songs

# songs = load_song_data()
# print(songs)


def pinyin_filter(text):
    try:
        # Get full Pinyin for the entire string
        result = pinyin(text, style=Style.NORMAL)
        return [item[0] for item in result]  # Return list of Pinyin syllables
    except:
        return []

# a = pinyin_filter("TEMPLATE")
# print(a)


# def sort_excel_sheets():
#     """Sort sheets in an Excel file alphabetically"""
#     # Load workbook
#     wb = load_workbook(SONG_DATABASE_PATH)
#
#     # Get sheet names and sort them
#     sheet_names = sorted(wb.sheetnames, key=pinyin_filter)
#
#     # Reorder sheets by creating a new workbook
#     new_wb = load_workbook(SONG_DATABASE_PATH, keep_vba=True)
#     new_wb._sheets = []  # Clear existing sheets
#
#     # Add sheets in sorted order
#     for name in sheet_names:
#         sheet = wb[name]
#         new_sheet = new_wb.create_sheet(title=name)
#         # Copy data and formatting
#         for row in sheet.iter_rows():
#             for cell in row:
#                 new_sheet[cell.coordinate].value = cell.value
#                 if cell.has_style:
#                     new_sheet[cell.coordinate]._style = cell._style
#
#     # Save the new workbook
#     new_wb.save(SORTED_SONG_DATABASE_PATH)
#     print(f"Sorted sheets saved to {SORTED_SONG_DATABASE_PATH}")
#
#
# # Usage
# sort_excel_sheets()
